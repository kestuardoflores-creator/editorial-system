"""
norm_excel.py — Conversor bidireccional JSON ↔ Excel para normativas.
Usado por el instalador (JSON→Excel) y el ensamblador (Excel→dict).
"""

from pathlib import Path
import json

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, Protection)
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "--quiet", "openpyxl"])
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, Protection)
    from openpyxl.utils import get_column_letter

# ── Column order shown in Excel ────────────────────────────────────────────────
STYLE_COLUMNS = [
    "ID_Etiqueta",
    "Nombre_Visible",
    "Fuente",
    "Tamano",
    "Interlineado",
    "Negrita",
    "Italica",
    "Sangria_1era",
    "Alineacion",
    "Color_Texto",
    "Espaciado_Antes",
    "Espaciado_Despues",
    "Es_Numerable",
    "Prefijo_Texto",
    "Separador_Num",
    "Formato_Prefijo",
    "Formato_Numero",
    "Posicion_Pagina",
    "Alineacion_Objeto",
]

ROOT_COLUMNS = ["normativa", "version", "inicio_capitulo",
                "chars_por_pagina"]

# ── Friendly column descriptions shown as a second header row ─────────────────
COLUMN_HINTS = {
    "ID_Etiqueta":      "ID (no editar)",
    "Nombre_Visible":   "Nombre visible",
    "Fuente":           "Familia tipográfica",
    "Tamano":           "Tamaño en puntos",
    "Interlineado":     "1.0 / 1.5 / 2.0",
    "Negrita":          "TRUE / FALSE",
    "Italica":          "TRUE / FALSE",
    "Sangria_1era":     "Puntos (0 = sin sangría)",
    "Alineacion":       "LEFT / CENTER / RIGHT / JUSTIFY",
    "Color_Texto":      "Hex (#000000)",
    "Espaciado_Antes":  "Puntos antes del párrafo",
    "Espaciado_Despues":"Puntos después del párrafo",
    "Es_Numerable":     "TRUE / FALSE",
    "Prefijo_Texto":    "Figura / Tabla / (vacío)",
    "Separador_Num":    ". o espacio",
    "Formato_Prefijo":  "CAPITULO_ELEMENTO / CONTINUO",
    "Formato_Numero":   "ARABIC / ROMAN_UPPER / ROMAN_LOWER",
    "Posicion_Pagina":  "BREAK_TEXT / INLINE",
    "Alineacion_Objeto":"CENTER / LEFT / RIGHT",
}

# ── Colors ─────────────────────────────────────────────────────────────────────
COLOR_HEADER = "1F4E79"   # dark blue
COLOR_HINT   = "2E75B6"   # medium blue
COLOR_ROOT   = "D6E4F0"   # light blue for root config row
COLOR_LOCKED = "F2F2F2"   # grey for non-editable cells
COLOR_ALT    = "EBF3FB"   # alternating row


def _header_font(size=10):
    return Font(name="Calibri", size=size, bold=True, color="FFFFFF")

def _hint_font():
    return Font(name="Calibri", size=9, italic=True, color="FFFFFF")

def _cell_font():
    return Font(name="Calibri", size=10)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────
# JSON → EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def json_to_excel(json_path, excel_path):
    """
    Convert a normativa JSON file to a styled Excel workbook.
    Creates two sheets: 'Estilos' and 'Configuracion'.
    """
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Estilos ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Estilos"
    ws.sheet_view.showGridLines = False

    # Row 1 — column headers
    for col_idx, col_name in enumerate(STYLE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = _header_font()
        cell.fill    = _fill(COLOR_HEADER)
        cell.border  = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    # Row 2 — hints
    for col_idx, col_name in enumerate(STYLE_COLUMNS, start=1):
        hint = COLUMN_HINTS.get(col_name, "")
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font    = _hint_font()
        cell.fill    = _fill(COLOR_HINT)
        cell.border  = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    # Lock ID_Etiqueta column header note
    ws.freeze_panes = "B3"

    # Data rows
    for row_idx, style in enumerate(data["estilos"], start=3):
        bg = COLOR_ALT if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, col_name in enumerate(STYLE_COLUMNS, start=1):
            value = style.get(col_name, "")
            # Booleans as uppercase strings for readability
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = _cell_font()
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Grey out ID column
            if col_name == "ID_Etiqueta":
                cell.fill = _fill(COLOR_LOCKED)
                cell.font = Font(name="Calibri", size=10,
                                 bold=True, color="1F4E79")
            else:
                cell.fill = _fill(bg)

    # Column widths
    col_widths = {
        "ID_Etiqueta": 18, "Nombre_Visible": 24, "Fuente": 18,
        "Tamano": 10, "Interlineado": 12, "Negrita": 10, "Italica": 10,
        "Sangria_1era": 14, "Alineacion": 14, "Color_Texto": 13,
        "Espaciado_Antes": 16, "Espaciado_Despues": 17,
        "Es_Numerable": 13, "Prefijo_Texto": 14, "Separador_Num": 12,
        "Formato_Prefijo": 22, "Formato_Numero": 18, "Posicion_Pagina": 16,
        "Alineacion_Objeto": 17,
    }
    for col_idx, col_name in enumerate(STYLE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # ── Sheet 2: Configuracion ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Configuracion")
    ws2.sheet_view.showGridLines = False

    # Header
    ws2.cell(row=1, column=1, value="Campo").font    = _header_font()
    ws2.cell(row=1, column=1).fill   = _fill(COLOR_HEADER)
    ws2.cell(row=1, column=1).border = _thin_border()
    ws2.cell(row=1, column=2, value="Valor").font    = _header_font()
    ws2.cell(row=1, column=2).fill   = _fill(COLOR_HEADER)
    ws2.cell(row=1, column=2).border = _thin_border()
    ws2.cell(row=1, column=3, value="Descripción").font = _header_font()
    ws2.cell(row=1, column=3).fill   = _fill(COLOR_HEADER)
    ws2.cell(row=1, column=3).border = _thin_border()

    cfg_rows = [
        ("normativa",        data.get("normativa", ""),
         "Nombre de la normativa"),
        ("version",          data.get("version", ""),
         "Versión del archivo"),
        ("inicio_capitulo",  data.get("inicio_capitulo", "IMPAR"),
         "IMPAR / PAR / NUEVA / CONTINUO"),
        ("chars_por_pagina", data.get("chars_por_pagina", 2500),
         "Caracteres estimados por página (para TOC)"),
        ("margen_top_cm",    data.get("margenes_cm", {}).get("top", 2.54),
         "Margen superior en cm"),
        ("margen_bottom_cm", data.get("margenes_cm", {}).get("bottom", 2.54),
         "Margen inferior en cm"),
        ("margen_left_cm",   data.get("margenes_cm", {}).get("left", 2.54),
         "Margen izquierdo en cm"),
        ("margen_right_cm",  data.get("margenes_cm", {}).get("right", 2.54),
         "Margen derecho en cm"),
    ]

    for i, (campo, valor, desc) in enumerate(cfg_rows, start=2):
        bg = COLOR_ALT if i % 2 == 0 else "FFFFFF"
        c1 = ws2.cell(row=i, column=1, value=campo)
        c1.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
        c1.fill = _fill(COLOR_LOCKED)
        c1.border = _thin_border()

        c2 = ws2.cell(row=i, column=2, value=valor)
        c2.font = _cell_font()
        c2.fill = _fill(bg)
        c2.border = _thin_border()

        c3 = ws2.cell(row=i, column=3, value=desc)
        c3.font = Font(name="Calibri", size=9, italic=True, color="595959")
        c3.fill = _fill(bg)
        c3.border = _thin_border()

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 40
    ws2.freeze_panes = "A2"

    wb.save(str(excel_path))


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL → DICT  (used by assembler at runtime)
# ─────────────────────────────────────────────────────────────────────────────

def excel_to_dict(excel_path):
    """
    Read a normativa Excel file and return a dict identical in structure
    to the original JSON, so the assembler needs no changes.
    """
    wb   = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws   = wb["Estilos"]
    ws2  = wb["Configuracion"]

    # Read Configuracion sheet
    cfg = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cfg[row[0]] = row[1]

    # Read Estilos — row 1 is headers, row 2 is hints, row 3+ is data
    headers = [cell.value for cell in ws[1]]
    estilos = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        style = {}
        for col_name, value in zip(headers, row):
            if value is None:
                value = ""
            # Convert TRUE/FALSE strings back to booleans
            if str(value).upper() == "TRUE":
                value = True
            elif str(value).upper() == "FALSE":
                value = False
            style[col_name] = value
        estilos.append(style)

    return {
        "normativa":      cfg.get("normativa", ""),
        "version":        cfg.get("version", ""),
        "inicio_capitulo":cfg.get("inicio_capitulo", "NUEVA"),
        "chars_por_pagina": cfg.get("chars_por_pagina", 2500),
        "margenes_cm": {
            "top":    cfg.get("margen_top_cm", 2.54),
            "bottom": cfg.get("margen_bottom_cm", 2.54),
            "left":   cfg.get("margen_left_cm", 2.54),
            "right":  cfg.get("margen_right_cm", 2.54),
        },
        "estilos": estilos,
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLI  —  python norm_excel.py <normativa>
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python norm_excel.py <nombre_normativa>")
        print("Ejemplo: python norm_excel.py apa7")
        sys.exit(1)

    name      = sys.argv[1].lower()
    base      = Path(__file__).parent.parent / "config"
    json_path = base / f"{name}.json"
    xlsx_path = base / f"{name}.xlsx"

    if not json_path.exists():
        print(f"❌ No se encontró: {json_path}")
        sys.exit(1)

    json_to_excel(json_path, xlsx_path)
    print(f"✅ Excel generado: {xlsx_path}")
